from pprint import pprint
import requests
from bs4 import BeautifulSoup
import gzip
import openpyxl
from lxml import etree

from .WebScraperAbstractController import WebScraperAbstractController


class WebScraperController(WebScraperAbstractController):
    def __init__(self, file_name=None):
        self.file_name = file_name
        self.wb = self.excel_create_or_load()
        super().__init__(
            file_name=file_name
        )

    def get_addresses_bestattungsvergleich_de(self):
        base_url = 'https://www.bestattungsvergleich.de/'
        html = self.fetch_page(base_url)
        soup = self.scraper(html, 'html.parser')
        states = {
            'Baden-Württemberg': 'verzeichnis/bestatter-in-baden-wuerttemberg',
            'Bayern': 'verzeichnis/bestatter-in-bayern',
            'Berlin': 'verzeichnis/bestatter-in-berlin',
            'Brandenburg': 'verzeichnis/bestatter-in-brandenburg',
            'Bremen': 'verzeichnis/bestatter-in-bremen',
            'Hamburg': 'verzeichnis/bestatter-in-hamburg',
            'Hessen': 'verzeichnis/bestatter-in-hessen',
            'Mecklenburg-Vorpommern': 'verzeichnis/bestatter-in-mecklenburg-vorpommern',
            'Niedersachsen': 'verzeichnis/bestatter-in-niedersachsen',
            'Nordrhein-Westfalen': 'verzeichnis/bestatter-in-nordrhein-westfalen',
            'Rheinland-Pfalz': 'verzeichnis/bestatter-in-rheinland-pfalz',
            'Saarland': 'verzeichnis/bestatter-in-saarland',
            'Sachsen': 'verzeichnis/bestatter-in-sachsen',
            'Sachsen-Anhalt': 'verzeichnis/bestatter-in-sachsen-anhalt',
            'Schleswig-Holstein': 'verzeichnis/bestatter-in-schleswig-holstein',
            'Thüringen': 'verzeichnis/bestatter-in-thueringen'
        }

        # Erstellen eines einziges Arbeitsblatts für alle Bundesländer
        all_states_sheet = self.wb.create_sheet(title="Alle Bundesländer")
        all_states_sheet.append(["Name", "Straße", "PLZ", "Ort", "Bundesland"])  # Spalte für Bundesland hinzugefügt

        for state_name, state_url in states.items():
            print(f"Scraping state: " + state_name)

            # Durchsuchen jeder Seite des Bundeslandes nach Adressen
            for page in range(1, 30):  # Beispiel: Seiten 1 bis 5
                page_url = f"https://www.bestattungsvergleich.de/{state_url}/seite/{page}"
                print("Scraping page: " + page_url)
                page_html = self.fetch_page(page_url)
                page_soup = self.scraper(page_html, 'html.parser')

                # Extrahieren der Adressdaten
                local_business_elements = page_soup.find_all(itemtype="http://schema.org/LocalBusiness")
                for element in local_business_elements:
                    name_element = element.find(itemprop="name")
                    if name_element:
                        name_anchor = name_element.find('a')  # find the first 'a' tag within the name_element
                        if name_anchor:
                            name = name_anchor.get_text(strip=True)  # get the text of the 'a' tag
                        else:
                            name = "Unbekannt"
                    else:
                        name = "Unbekannt"
                    address_element = element.find(itemtype="http://schema.org/PostalAddress")

                    if address_element:
                        street = address_element.find(itemprop="streetAddress").get_text(
                            strip=True) if address_element.find(itemprop="streetAddress") else ""
                        postal_code = address_element.find(itemprop="postalCode").get_text(
                            strip=True) if address_element.find(itemprop="postalCode") else ""
                        city = address_element.find(itemprop="addressLocality").get_text(
                            strip=True) if address_element.find(itemprop="addressLocality") else ""

                        print(name, street, address_element.find(itemprop="postalCode").get_text(
                            strip=True), city)

                        # Hinzufügen der Daten zum Excel-Sheet
                        all_states_sheet.append([name, street, str(postal_code), city,
                                                 state_name])  # state_name für die Bundeslandspalte hinzugefügt

        # Speichern des Excel-Workbooks
        self.excel_save_workbook()

    def get_addresses_bestatter_at(self):
        base_url = 'https://www.bestatter.at'

        states = {
            "Burgenland": "/ihre-bestatter/bundesland/burgenland",
            "Kärnten": "/ihre-bestatter/bundesland/karnten",
            "Niederösterreich": "/ihre-bestatter/bundesland/niederosterreich",
            "Oberösterreich": "/ihre-bestatter/bundesland/oberosterreich",
            "Salzburg": "/ihre-bestatter/bundesland/salzburg",
            "Steiermark": "/ihre-bestatter/bundesland/steiermark",
            "Tirol": "/ihre-bestatter/bundesland/tirol",
            "Vorarlberg": "/ihre-bestatter/bundesland/vorarlberg",
            "Wien": "/ihre-bestatter/bundesland/wien"
        }

        # Erstellen eines einziges Arbeitsblatts für alle Bundesländer
        all_states_sheet = self.wb.create_sheet(title="Alle Bundesländer")
        all_states_sheet.append(["Name", "Adresse", "Bundesland"])  # Spalte für Bundesland hinzugefügt

        """
        Bundesländer
        """
        for state_name, state_url in states.items():
            print(f"Scraping state: " + state_name)

            state_html = self.fetch_page(base_url + state_url)
            state_soup = self.scraper(state_html, 'html.parser')

            pages_elements = state_soup.select('nav.pagination li.last a.last')
            if pages_elements:
                pages_element = pages_elements[0]
                pages = pages_element.get("href").split('/')[-1]
                print(f"Bundesland {state_name} has {pages} pages")
            else:
                pages_elements = state_soup.select('nav.pagination li a')
                pages = len(pages_elements) if pages_elements else 0

            """
            Pages
            """
            pages = int(pages)
            if pages > 0:
                pages = int(pages)
                for page_number in range(1, pages + 1):
                    print(f"Scraping Page {page_number}/{pages+1}")
                    state_page_url = f"{base_url}{state_url}/page_mmce33/{page_number}"
                    state_page_html = self.fetch_page(state_page_url)
                    state_page_soup = self.scraper(state_page_html, 'html.parser')

                    companies_urls = []
                    for a_tag in state_page_soup.find_all('ul', class_='bestatter_liste')[0].find_all('li'):
                        a = a_tag.find('a')
                        if a:  # Stelle sicher, dass ein `<a>`-Tag gefunden wurde
                            companies_urls.append(a.get('href'))

                    """
                    Company Site
                    """
                    for company_url in companies_urls:
                        # Company Site
                        company_html = self.fetch_page(base_url + company_url)
                        company_soup = self.scraper(company_html, 'html.parser')

                        try:
                            name = company_soup.select("div.ce_metamodel_content.first h1")[0].get_text()
                            address = company_soup.select("div.unternehmen-detail p")[0].contents[0].strip()
                            # Hinzufügen der Daten zum Excel-Sheet
                            all_states_sheet.append([name, address, state_name])  # state_name für die Bundeslandspalte hinzugefügt
                            # print(f"Firma {name} in der {address} - {company_url}")
                        except Exception as e:
                            print(f"No Contact found in url {company_url}, error: {e}")
                            continue

            else:
                print(f"No pages for {state_name}. Skip to the next Bundesland")
                continue


        # Speichern des Excel-Workbooks
        self.excel_save_workbook()

    def get_addresses_bdb(self):
        base_url = 'https://bdb-qr.de/32_Betreuer_alphabetisch.php?von=a&bis=z'

        # Erstellen eines einziges Arbeitsblatts für alle Bundesländer
        sheet = self.wb.create_sheet(title="Betreuer A-Z")
        sheet.append(["Firma", "Name", "Strasse", "PLZ Ort"])  # Spalte für Bundesland hinzugefügt

        html = self.fetch_page(base_url)
        html_soup = self.scraper(html, 'html.parser')

        info_elements = html_soup.findAll("div", class_="betreuer_kurzinfotext")

        for info in info_elements:
            company = info.select('h3')[0].contents[0].strip()
            additional_info_parts = [part.strip() for part in info.select('h3')[0].contents[2:] if isinstance(part, str)]
            additional_info = " | ".join(additional_info_parts)
            street = info.select("p")[0].contents[0].strip()
            plz_ort = info.select("p")[0].contents[2].strip()
            if company and street and plz_ort:
                sheet.append([company, additional_info, street, plz_ort])

        # Speichern des Excel-Workbooks
        self.excel_save_workbook()


    def get_addresses_evangelische_kirchen_bayern(self):
        """
        Diese Methode gibt alle Adressen evangelischer Kirchen in Bayern zurück.
        """
        dekanate_url = "https://landeskirche.bayern-evangelisch.de/service.php?o=dekanatmap&f=getDekanatMapData&r=json"
        dekanate_response = self.requests.get(dekanate_url)
        dekanate_response.raise_for_status()

        dekanate_list = [{
            "dekCode": dekanat_info["dekCode"],
            "elkbid": dekanat_info["elkbid"]
        } for dekanat_info in dekanate_response.json()["paths"].values()]

        kirchen_sheet = self.excel_add_sheet("Bayern")

        for dekanat in dekanate_list:
            gemeinde_url = f"https://landeskirche.bayern-evangelisch.de/service.php?o=dekanatmap&f=getDekanatInfo&dl=user&dekCode={dekanat['dekCode']}&elkbid={dekanat['elkbid']}&r=json"
            print(gemeinde_url)
            response = self.requests.get(gemeinde_url)
            response.raise_for_status()

            for gemeinde in response.json()["content"]["gemeinden"]:
                title = gemeinde.get('title')
                street = gemeinde.get('street')
                plz = gemeinde.get('pcode')
                city = gemeinde.get('locality')

                kirchen_sheet.append([title, street, plz, city])

        self.excel_save_workbook()
        return True

    def get_addresses_katholische_kirchen_munic(self):
        soup_links = self.parse_html("https://www.erzbistum-muenchen.de/pfarrei/linkliste-pfarreien/84000")
        link_elements = soup_links.find_all('a', class_="web")
        for link_element in link_elements:
            link = link_element.get('href')
            soup_pfarrei = self.parse_html(link)
            pfarrei_info_element = soup_pfarrei.find('h2', class_="orga-info")

    def get_prices_mappei(self):
        # 1 Get sitemap from url "https://shop.mappei.de/de/sitemap.xml"
        response_sitemap_general = requests.get("https://shop.mappei.de/de/sitemap.xml")
        sitemap_general_xml = response_sitemap_general.text

        # 2 parse the sitemap.xml to get the content of loc
        soup = BeautifulSoup(sitemap_general_xml, 'lxml-xml')
        loc_content = soup.find("loc").string

        # 3 get the gz file from loc and unzip it
        response_gz = requests.get(loc_content)
        gz_file_data = gzip.decompress(response_gz.content)

        # Parse the XML data
        soup = BeautifulSoup(gz_file_data, 'lxml-xml')

        # Find all <loc> elements
        locs = soup.find_all('loc')

        product_urls = []

        # Define a limit for number of products
        limit = 10
        count = 0

        # Loop through each <loc> element
        for loc in locs:
            url = loc.string
            response = requests.get(url)
            webpage = response.text
            page_soup = BeautifulSoup(webpage, 'lxml')
            parsed_html = etree.HTML(str(page_soup))

            # Check if the .product-detail class is in the HTML
            if page_soup.select_one('.product-detail'):
                product_urls.append(url)
                count += 1

                # You can call get_itemprop here for each itemprop you want to get.
                # For example, to get 'name' itemprop:
                sku = self.get_itemprop(parsed_html, 'sku')
                name = self.get_itemprop(parsed_html,  'name')
                currency = self.get_itemprop_price(parsed_html, 'priceCurrency')
                price_low, price_high = self.get_itemprop_price(parsed_html, 'price', 'content')
                price_low_formatted = f'{float(price_low):.2f}'.replace('.', ',') if price_low and price_low.strip() else None
                price_high_formatted = f'{float(price_high):.2f}'.replace('.', ',') if price_high and price_high.strip() else None

                print(f'"{sku}";"";"{price_high_formatted}"')

        pprint(product_urls)
        return gz_file_data

    @staticmethod
    def get_itemprop(content, itemproperty, attribute=None, nr=0):
        """
        Get the value of a specific itemprop from the content. If attribute is specified, get the value of the attribute.

        Parameters:
        content (object): Parsed HTML document.
        itemproperty (str): The itemprop to be retrieved.
        attribute (str, optional): The attribute of itemprop to be retrieved. Defaults to None.
        nr (int, optional): The index of the itemprop's value if there are more than one. Defaults to 0.

        Returns:
        str: The value of the itemprop as a string, or the value of the attribute of the itemprop.
        """
        try:
            # Retrieving element containing itemprop, ignores the tag
            xpath_query = f'//*[contains(@class, "product-detail")]//*[@itemprop="{itemproperty}"]'
            element = content.xpath(xpath_query)
            if element:
                # if attribute is specified, return the value of the attribute
                if attribute:
                    return element[nr].attrib.get(attribute, None)
                else:
                    return ''.join(element[nr].xpath('.//text()')).strip()
            else:
                return None
        except Exception as e:
            print('An error occurred while retrieving the value of the itemprop: {}'.format(e))
            return None

    @staticmethod
    def get_itemprop_price(content, itemproperty, attribute=None, nr=0):
        """
        ...
        """
        try:
            xpath_query = f'//*[contains(@class, "product-detail")]//*[@itemprop="offers"]//*[@itemprop="{itemproperty}"]'
            element = content.xpath(xpath_query)
            price = element[nr].attrib.get(attribute, None) if element and attribute else ''.join(
                element[nr].xpath('.//text()')).strip() if element else None

            offer_count_element = content.xpath(f'//*[contains(@class, "product-detail")]//*[@itemprop="offerCount"]')
            if offer_count_element and int(offer_count_element[0].attrib.get('content', '0')) > 1:
                xpath_query_low = f'//*[contains(@class, "product-detail")]//*[@itemprop="lowPrice"]'
                xpath_query_high = f'//*[contains(@class, "product-detail")]//*[@itemprop="highPrice"]'
                low_price_elm = content.xpath(xpath_query_low)
                high_price_elm = content.xpath(xpath_query_high)
                if low_price_elm and high_price_elm:
                    if attribute:
                        low_price = low_price_elm[nr].attrib.get(attribute, None)
                        high_price = high_price_elm[nr].attrib.get(attribute, None)
                    else:
                        low_price = ''.join(low_price_elm[nr].xpath('.//text()')).strip()
                        high_price = ''.join(high_price_elm[nr].xpath('.//text()')).strip()
                else:
                    low_price = high_price = price
            else:
                low_price = high_price = price

            return low_price, high_price

        except Exception as e:
            print('An error occurred while retrieving the value of the itemprop: {}'.format(e))
            return None, None





