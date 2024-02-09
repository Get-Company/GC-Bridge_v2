from pprint import pprint

import openpyxl
import requests
import os

from ..ModulesCoreController import ModulesCoreController
from bs4 import BeautifulSoup


class WebScraperCoreController(ModulesCoreController):
    def __init__(self, file_name):

        self.scraper = BeautifulSoup
        self.requests = requests
        self.url = None
        self.response = None
        self.file_name = file_name
        self.wb = self.excel_create_or_load()

        super().__init__()

    def set_url(self, url):
        self.url = url

    def fetch_page(self,url=None):
        if url:
            self.set_url(url=url)
        self.response = self.requests.get(self.url)
        self.response.raise_for_status()
        return self.response.text

    def parse_html(self, url):
        html_text = self.fetch_page(url=url)
        return BeautifulSoup(html_text, 'html.parser')

    def normalize_address(self, name, street, plz, city):
        """ created by rocket_ln3 19.01.2024 """
        address = f"{name}, {street}, {plz} {city}"
        return address

    def excel_create_or_load(self):
        if os.path.exists(self.file_name):
            wb = openpyxl.load_workbook(self.file_name)
        else:
            wb = openpyxl.Workbook()

        return wb

    def excel_add_sheet(self, sheet_name):
        sheet = self.wb.create_sheet(sheet_name)
        headers = ["Name", "Street", "ZIP", "City"]
        sheet.append(headers)
        return sheet

    def excel_add_or_get_sheet(self, sheet_name):
        # Überprüfen, ob das Tabellenblatt bereits existiert
        if sheet_name in self.wb.sheetnames:
            # Wenn das Tabellenblatt existiert, es auswählen
            sheet = self.wb[sheet_name]
        else:
            # Wenn das Tabellenblatt nicht existiert, ein neues erstellen
            sheet = self.wb.create_sheet(sheet_name)
            # Überschriften für das neue Tabellenblatt hinzufügen
            headers = ["Name", "Street", "ZIP", "City"]
            sheet.append(headers)

        return sheet

    def excel_remove_file(self, filename):
        if os.path.exists(filename):
            os.remove(filename)
            print(f"File {filename} was deletetd succesfully")
            return True
        else:
            print(f"Could not remove file {filename}")
            return False

    def excel_remove_sheet(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            sheet = self.wb[sheet_name]
            self.wb.remove(sheet)

    def excel_save_workbook(self):
        self.wb.save(self.file_name)

    """
    Other methods not needed
    """

    def sync_all_to_bridge(self):
        pass

    def sync_all_from_bridge(self, bridge_entities):
        pass

    def sync_one_to_bridge(self):
        pass

    def sync_one_from_bridge(self, bridge_entity):
        pass

    def sync_changed_to_bridge(self):
        pass

    def sync_changed_from_bridge(self, bridge_entities):
        pass

